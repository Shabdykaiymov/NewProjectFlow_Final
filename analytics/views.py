from rest_framework import views, status, permissions
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.db.models import Count
from projects.models import Project
from tasks.models import Task
from .services import AnalyticsService
from openpyxl import Workbook
import datetime


class TasksByStatusView(views.APIView):
    """
    Представление для получения статистики задач по статусам
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Получаем ID проекта из запроса, если он есть
        project_id = request.query_params.get('project_id')

        # Получаем статистику задач по статусам
        data = AnalyticsService.get_tasks_by_status(request.user, project_id)
        return Response(data)


class TasksByUserView(views.APIView):
    """
    Представление для получения статистики задач по пользователям
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Получаем ID проекта из запроса, если он есть
        project_id = request.query_params.get('project_id')

        # Если указан ID проекта, проверяем права доступа
        if project_id:
            try:
                project = Project.objects.get(id=project_id)
                if not project.members.filter(id=request.user.id).exists():
                    return Response(
                        {'error': 'У вас нет доступа к этому проекту'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except Project.DoesNotExist:
                return Response(
                    {'error': 'Проект не найден'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Получаем статистику задач по пользователям для проекта
            data = AnalyticsService.get_tasks_by_user(project_id=project_id)
        else:
            # Получаем статистику задач по пользователям для всех проектов
            data = AnalyticsService.get_tasks_by_user(user=request.user)

        return Response(data)


class ProjectProgressView(views.APIView):
    """
    Представление для получения прогресса по проекту
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, project_id):
        # Проверяем права доступа к проекту
        project = get_object_or_404(Project, id=project_id)
        if not project.members.filter(id=request.user.id).exists():
            return Response(
                {'error': 'У вас нет доступа к этому проекту'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Получаем данные о прогрессе проекта
        data = AnalyticsService.get_project_progress(project_id)
        return Response(data)


class ExportReportView(views.APIView):
    """
    Представление для экспорта отчета в Excel
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # Получаем ID проекта из запроса, если он есть
        project_id = request.query_params.get('project_id')

        # Создаем новый workbook
        wb = Workbook()
        ws = wb.active

        # Устанавливаем название листа
        ws.title = "Отчет по задачам"

        # Получаем данные
        user = request.user
        projects = Project.objects.filter(members=user)

        # Базовый запрос
        tasks_query = Task.objects.filter(project__in=projects)

        # Добавляем название проекта в первую строку
        if project_id:
            try:
                project = Project.objects.get(id=project_id)
                if not project.members.filter(id=user.id).exists():
                    return Response(
                        {'error': 'У вас нет доступа к этому проекту'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                tasks_query = tasks_query.filter(project_id=project_id)
                ws.cell(row=1, column=1, value=f"Проект: {project.name}")
            except Project.DoesNotExist:
                return Response(
                    {'error': 'Проект не найден'},
                    status=status.HTTP_404_NOT_FOUND
                )
        else:
            ws.cell(row=1, column=1, value="Все проекты")

        # Добавляем пустую строку
        start_row = 3

        # Заголовки столбцов на строке 3
        headers = ['Статус', 'Количество задач']
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)

        # Подсчитываем задачи по статусам
        status_data = tasks_query.values('status').annotate(count=Count('id'))

        # Заполняем данные начиная со строки 4
        row = start_row + 1
        total_tasks = 0

        # Все возможные статусы
        for status_key, status_name in Task.STATUS_CHOICES:
            # Ищем количество для данного статуса
            count = 0
            for item in status_data:
                if item['status'] == status_key:
                    count = item['count']
                    break

            ws.cell(row=row, column=1, value=status_name)
            ws.cell(row=row, column=2, value=count)
            total_tasks += count
            row += 1

        # Добавляем итоговую строку
        ws.cell(row=row + 1, column=1, value="ИТОГО:")
        ws.cell(row=row + 1, column=2, value=total_tasks)

        # Настраиваем ширину столбцов
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        # Делаем жирным заголовок проекта
        from openpyxl.styles import Font
        ws.cell(row=1, column=1).font = Font(bold=True)

        # Делаем жирными заголовки столбцов
        for col in range(1, 3):
            ws.cell(row=start_row, column=col).font = Font(bold=True)

        # Делаем жирной итоговую строку
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        ws.cell(row=row + 1, column=2).font = Font(bold=True)

        # Создаем HTTP ответ с Excel файлом
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'tasks_report_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        # Сохраняем workbook в response
        wb.save(response)
        return response